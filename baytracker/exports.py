"""
exports.py -- build the four purpose-shaped tables and write them to CSV/XLSX.

We deliberately do NOT export the raw event log as the primary artifact. Instead
we derive four tables where ONE ROW = ONE REAL THING (spec section 8):

  1. Delays        -- one row per delay episode
  2. Bay Runs      -- one row per time a unit occupied a bay
  3. Unit Journeys -- one row per work order
  4. Events (raw)  -- the full append-only log (the source of truth)

The same derivation feeds the Stats page charts, so the numbers on screen and in
the export always agree. Honesty rules (Appendix C2): open runs/delays show blank
end times and blank durations -- never a fabricated end -- and cost only appears
when a real labor rate has been entered.

Formatting conventions (spec section 8):
  * Timestamps ISO 8601 "YYYY-MM-DD HH:MM:SS" (seconds kept for precision).
  * Durations as plain decimal minutes (sums/averages cleanly) + an H:MM twin.
  * work_order / product_number written as TEXT so Excel can't eat leading zeros.
  * UTF-8 with BOM; commas/newlines in notes are properly quoted.
  * Column order: identifiers -> times -> durations -> categories -> notes/initials.
"""

from __future__ import annotations

import csv
import io
import zipfile
from datetime import datetime
from typing import Dict, List, Optional

from . import db, events, state
from .schedule import Schedule


# ---------------------------------------------------------------------------
# Small formatting helpers
# ---------------------------------------------------------------------------

def _minutes(seconds: Optional[float]) -> Optional[float]:
    """Seconds -> decimal minutes rounded to 0.1, or None (kept blank)."""
    if seconds is None:
        return None
    return round(seconds / 60.0, 1)


def _hmm(seconds: Optional[float]) -> Optional[str]:
    """Seconds -> 'H:MM' (hours may exceed 24), or None."""
    if seconds is None:
        return None
    total = int(round(seconds))
    h = total // 3600
    m = (total % 3600) // 60
    return f"{h}:{m:02d}"


def _bay_lookups(conn):
    rows = conn.execute("SELECT id, name, sort_order FROM bays;").fetchall()
    name = {r["id"]: r["name"] for r in rows}
    number = {r["id"]: r["sort_order"] for r in rows}
    return name, number


# ---------------------------------------------------------------------------
# Derivation: turn the replay into the four lists of rows (UNFILTERED).
# Each row carries a hidden '_dt' (a datetime) used for date-range/shift
# filtering; hidden keys are stripped before output.
# ---------------------------------------------------------------------------

def derive_rows(conn, now: Optional[datetime] = None) -> Dict[str, List[dict]]:
    now = now or datetime.now()
    sched = Schedule.from_settings(conn)
    r = state.replay(conn)
    bay_name, bay_number = _bay_lookups(conn)
    labor_rate = db.get_setting(conn, "labor_rate", None)

    # ---- Delays ----
    delays = []
    for d in r.delays:
        closed = d.cleared is not None
        secs = sched.counted_seconds(d.started, d.cleared) if closed else None
        mins = _minutes(secs)
        est_cost = None
        if closed and labor_rate not in (None, "", 0):
            try:
                est_cost = round((secs / 3600.0) * float(labor_rate), 2)
            except (TypeError, ValueError):
                est_cost = None
        delays.append({
            "_dt": d.started,
            "_division": d.division,
            "_reason": d.reason_label,
            "_product": d.product_number,
            "_bay_id": d.bay_id,
            "work_order": d.work_order,
            "product_number": d.product_number,
            "bay": bay_name.get(d.bay_id, d.bay_id),
            "started_at": _ts(d.started),
            "cleared_at": _ts(d.cleared),               # blank if open
            "delay_minutes": mins,                        # blank if open
            "delay_hmm": _hmm(secs),
            "reason": d.reason_label,
            "division": d.division,
            "in_or_out_of_control": d.in_out_of_control,
            "shift": sched.shift_for(d.started),
            "est_cost": est_cost,
            "note": d.note,
            "flagged_by": d.flagged_by,
            "cleared_by": d.cleared_by,
        })

    # ---- Bay Runs ----
    bay_runs = []
    for run in r.runs:
        end = run.ended
        if end is not None:
            active = state.counted_over(sched, run.running_intervals(end))
            delay = state.counted_over(sched, run.delayed_intervals(end))
            total = active + delay
        else:
            active = delay = total = None  # open run -> blank durations
        bay_runs.append({
            "_dt": run.started,
            "_product": run.product_number,
            "_bay_id": run.bay_id,
            "work_order": run.work_order,
            "product_number": run.product_number,
            "bay": bay_name.get(run.bay_id, run.bay_id),
            "started_at": _ts(run.started),
            "ended_at": _ts(run.ended),                  # blank if open
            "active_minutes": _minutes(active),
            "delay_minutes": _minutes(delay),
            "total_minutes": _minutes(total),
            "active_hmm": _hmm(active),
            "shift": sched.shift_for(run.started),
            "started_by": run.started_by,
            "completed_by": run.completed_by,
        })

    # ---- Unit Journeys ----
    unit_journeys = []
    for wo, u in r.units.items():
        dur = state.unit_durations(sched, u, r.runs, now)
        path = " → ".join(str(bay_number.get(b, b)) for b in u.bays_visited)
        unit_journeys.append({
            "_dt": u.first_started,
            "_product": u.product_number,
            "work_order": wo,
            "product_number": u.product_number,
            "first_started_at": _ts(u.first_started),
            "completed_at": _ts(u.completed),            # blank if open
            "cycle_minutes": _minutes(dur["cycle"]),
            "active_minutes": _minutes(dur["active"]),
            "delay_minutes": _minutes(dur["delay"]),
            "queue_minutes": _minutes(dur["queue"]),
            "bays_visited": path,
            "delay_count": u.delay_count,
            "outcome": u.outcome or "open",
        })

    # ---- Events (raw) ----
    raw = []
    for e in events.all_events(conn):
        d = dict(e)
        d["_dt"] = events.parse_ts(e["ts"])
        raw.append(d)

    return {"delays": delays, "bay_runs": bay_runs,
            "unit_journeys": unit_journeys, "events": raw}


def _ts(dt: Optional[datetime]) -> Optional[str]:
    from . import config
    return dt.strftime(config.TS_FORMAT) if dt else None


# ---------------------------------------------------------------------------
# Filtering. "export everything" simply passes an empty filter dict.
# ---------------------------------------------------------------------------

def apply_filters(rows: List[dict], filters: dict, table: str) -> List[dict]:
    start = filters.get("start")   # ISO date or datetime string, inclusive
    end = filters.get("end")
    bay_id = filters.get("bay_id")
    reason = filters.get("reason")
    division = filters.get("division")
    product = filters.get("product_number")
    shift = filters.get("shift")

    def keep(row) -> bool:
        dt: datetime = row["_dt"]
        if start and dt < _parse_filter_dt(start, end_of_day=False):
            return False
        if end and dt > _parse_filter_dt(end, end_of_day=True):
            return False
        if bay_id and row.get("_bay_id") not in (None, int(bay_id)):
            return False
        # reason/division apply to the Delays table; product/shift where present.
        if reason and row.get("_reason") not in (None, reason):
            if "reason" in row and row.get("reason") != reason:
                return False
        if division and row.get("division") not in (None, division):
            return False
        if product and row.get("_product") not in (None, product) and row.get("product_number") != product:
            return False
        if shift and row.get("shift") not in (None, shift):
            return False
        return True

    return [r for r in rows if keep(r)]


def _parse_filter_dt(value: str, end_of_day: bool) -> datetime:
    """Accept 'YYYY-MM-DD' or full timestamps in a filter."""
    value = value.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    # Date only:
    d = datetime.strptime(value, "%Y-%m-%d")
    return d.replace(hour=23, minute=59, second=59) if end_of_day else d


def _strip_hidden(row: dict) -> dict:
    return {k: v for k, v in row.items() if not k.startswith("_")}


# ---------------------------------------------------------------------------
# Column order per table (identifiers -> times -> durations -> categories ->
# notes/initials), and human-readable Data Dictionary text.
# ---------------------------------------------------------------------------

COLUMNS = {
    "delays": ["work_order", "product_number", "bay", "started_at", "cleared_at",
               "delay_minutes", "delay_hmm", "reason", "division",
               "in_or_out_of_control", "shift", "est_cost", "note",
               "flagged_by", "cleared_by"],
    "bay_runs": ["work_order", "product_number", "bay", "started_at", "ended_at",
                 "active_minutes", "delay_minutes", "total_minutes", "active_hmm",
                 "shift", "started_by", "completed_by"],
    "unit_journeys": ["work_order", "product_number", "first_started_at",
                      "completed_at", "cycle_minutes", "active_minutes",
                      "delay_minutes", "queue_minutes", "bays_visited",
                      "delay_count", "outcome"],
    "events": ["id"] + events.EVENT_COLUMNS,
}

# Columns that must be forced to TEXT in XLSX so Excel cannot reinterpret them
# (leading zeros / scientific notation).
TEXT_COLUMNS = {"work_order", "product_number"}

SHEET_TITLES = {
    "delays": "Delays",
    "bay_runs": "Bay Runs",
    "unit_journeys": "Unit Journeys",
    "events": "Events",
}

DATA_DICTIONARY = [
    ("ABOUT", "How the minute columns are computed", ""),
    ("", "Non-counting time", "Scheduled breaks AND off-hours (time outside the operating "
        "calendar) never count. All minute columns below have these removed."),
    ("", "active_minutes", "Time a bay was RUNNING: occupied time minus delay time minus "
        "non-counting time."),
    ("", "delay_minutes", "Time a bay was DELAYED, minus non-counting time. At the unit "
        "level, delay only counts when no other bay of that unit was running."),
    ("", "queue_minutes", "Time a unit sat in the WIP pool (occupying zero bays) before "
        "being started elsewhere or completed, minus non-counting time."),
    ("", "cycle_minutes", "active + delay + queue for the whole unit. Parallel work in two "
        "bays counts the UNION of active periods, never the sum."),
    ("", "Open records", "A run or delay that was never closed shows a blank end time and "
        "blank minutes -- it is never assigned a made-up end (Appendix C2)."),
    ("", "est_cost", "delay hours x the labor rate entered in Admin. Blank if no rate is set."),
    ("Delays", "work_order", "The unit's unique id (text -- leading zeros preserved)."),
    ("Delays", "product_number", "The unit's type/variation (text)."),
    ("Delays", "bay", "Bay the delay occurred in."),
    ("Delays", "started_at / cleared_at", "When the delay was flagged / cleared (ISO 8601)."),
    ("Delays", "delay_minutes / delay_hmm", "Counted delay duration as decimal minutes / H:MM."),
    ("Delays", "reason / division / in_or_out_of_control", "Snapshot of the delay reason and "
        "its owning division + control tag, captured at flag time."),
    ("Delays", "shift", "Shift the delay's start time falls in (by configured cutoffs)."),
    ("Delays", "note / flagged_by / cleared_by", "Required note, and who flagged/cleared it."),
    ("Bay Runs", "started_at / ended_at", "When the unit entered / left this bay."),
    ("Bay Runs", "active_minutes / delay_minutes / total_minutes", "Running / delayed / "
        "occupied counted minutes for this one bay occupancy."),
    ("Unit Journeys", "first_started_at / completed_at", "First START to terminal completion."),
    ("Unit Journeys", "bays_visited", "Order of bays the unit passed through, e.g. 3 -> 7 -> 5."),
    ("Unit Journeys", "delay_count", "Number of delays flagged across the whole journey."),
    ("Unit Journeys", "outcome", "complete, scrap, or open (still in progress)."),
    ("Events", "(all columns)", "The raw append-only log -- the source of truth from which "
        "every other table is derived. Never edited; corrections are new CORRECTION rows."),
]


# ---------------------------------------------------------------------------
# Filenames
# ---------------------------------------------------------------------------

def filename_suffix(filters: dict) -> str:
    """Self-documenting filename suffix from the active date filter."""
    start, end = filters.get("start"), filters.get("end")
    if start and end:
        return f"{start[:10]}_to_{end[:10]}"
    if start:
        return f"from_{start[:10]}"
    if end:
        return f"through_{end[:10]}"
    return "all"


# ---------------------------------------------------------------------------
# CSV (zip of 4 files + README) and XLSX (one workbook, 5 tabs)
# ---------------------------------------------------------------------------

def build_filtered(conn, filters: dict, now: Optional[datetime] = None) -> Dict[str, List[dict]]:
    derived = derive_rows(conn, now)
    out = {}
    for table, rows in derived.items():
        filtered = apply_filters(rows, filters, table)
        out[table] = [_strip_hidden(r) for r in filtered]
    return out


def to_csv_zip(conn, filters: dict, now: Optional[datetime] = None) -> bytes:
    """Return a .zip containing four BOM-UTF-8 CSVs + a README data dictionary."""
    tables = build_filtered(conn, filters, now)
    suffix = filename_suffix(filters)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for table, rows in tables.items():
            csv_text = _rows_to_csv(table, rows)
            # UTF-8 BOM so Excel opens it with the right encoding.
            zf.writestr(f"{table}_{suffix}.csv", "﻿" + csv_text)
        zf.writestr("README_data_dictionary.txt", _data_dictionary_text())
    return buf.getvalue()


def _rows_to_csv(table: str, rows: List[dict]) -> str:
    cols = COLUMNS[table]
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=cols, extrasaction="ignore",
                            lineterminator="\r\n")
    writer.writeheader()
    for r in rows:
        writer.writerow({c: ("" if r.get(c) is None else r.get(c)) for c in cols})
    return out.getvalue()


def _data_dictionary_text() -> str:
    lines = ["BAY TRACKING SYSTEM -- DATA DICTIONARY", "=" * 40, ""]
    current = None
    for table, col, desc in DATA_DICTIONARY:
        if table and table != current:
            current = table
            lines.append("")
            lines.append(f"[{table}]")
        prefix = f"  {col}: " if col else "  "
        lines.append(f"{prefix}{desc}")
    return "\r\n".join(lines) + "\r\n"


def to_xlsx(conn, filters: dict, now: Optional[datetime] = None) -> bytes:
    """Return an .xlsx workbook: 4 data tabs + a Data Dictionary tab.

    work_order and product_number cells are forced to text so Excel does not
    strip leading zeros or switch to scientific notation -- this is why XLSX is
    the safer human-readable artifact.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    tables = build_filtered(conn, filters, now)
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    for table, rows in tables.items():
        ws = wb.create_sheet(title=SHEET_TITLES.get(table, table))
        cols = COLUMNS[table]
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([_xlsx_value(c, r.get(c)) for c in cols])
        # Force identifier columns to text format.
        for idx, col in enumerate(cols, start=1):
            if col in TEXT_COLUMNS:
                for row_cells in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    for cell in row_cells:
                        cell.number_format = "@"  # Excel text format
        ws.freeze_panes = "A2"

    # Data Dictionary tab
    ws = wb.create_sheet(title="Data Dictionary")
    ws.append(["Table", "Column", "Definition"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for table, col, desc in DATA_DICTIONARY:
        ws.append([table, col, desc])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_value(col: str, value):
    if value is None:
        return ""
    if col in TEXT_COLUMNS:
        return str(value)   # keep as text (paired with the "@" number_format)
    return value
